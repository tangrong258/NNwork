import numpy as np
from openpyxl import load_workbook
import pandas as pd
from openpyxl import Workbook
import os
import xlsxwriter


# for i in range(0,3):
#     a = np.zeros((5,5))
#     workbook = Workbook()
#     track = r"C:\Users\tangrong\Desktop\conlution.xlsx"
#     workbook.save(track)
#     Writer = pd.ExcelWriter(track, engine='openpyxl')
#     book = load_workbook(Writer.path)
#     Writer.book = book
#     dt1 = pd.DataFrame(a)
#     dt1.to_excel(Writer, sheet_name=str(i))
#     Writer.save()
#     Writer.close()


xls2 = xlsxwriter.Workbook(r'D:\轨迹预测\conlution.xlsx')
sheet = xls2.add_worksheet('sheet1')
# sheet.write(0,0,'第一行第一列')
# sheet.write(0,1,'第一行第二列')
xls2.close()


for i in range(0,3):
    a = np.zeros((5,5))
    # workbook = Workbook()
    track = r'D:\轨迹预测\conlution.xlsx'
    # workbook.save(track)
    Writer = pd.ExcelWriter(track, engine='openpyxl')
    book = load_workbook(Writer.path)
    Writer.book = book
    dt1 = pd.DataFrame(a)
    dt1.to_excel(Writer, sheet_name=str(i))
    Writer.save()