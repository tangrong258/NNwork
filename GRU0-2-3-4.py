import tensorflow as tf
import pickle
import numpy as np
import os
import tensorflow.contrib as contrib
import matplotlib.pyplot as plt
from sklearn.preprocessing import MinMaxScaler
import pandas as pd
from tensorflow.python.keras.backend import set_session
from openpyxl import load_workbook
from openpyxl import Workbook

howlong = "60min"
sp = 60
rootfrom = r"D:\ZSNJAP01\flight\delaydata" + "\\"+ howlong
root = r"D:\ZSNJAP01\flight\prediction\GRU" + "\\" + howlong
sub = "lineavgt"
track = root + "\\" + sub + "\\" + "error" + "\\" + "conlution.xlsx"
workbook = Workbook()
workbook.save(track)

with open(rootfrom + "\\" + "lineavgt.pkl",'rb') as f:
    featime = pickle.load(f)
featime2d = np.reshape(featime,[-1,featime.shape[1] * featime.shape[2]])
flytime = np.zeros((featime2d.shape[0], 6), dtype='float32')
flytime[:, 0] = featime2d[:, 2]
flytime[:, 1] = featime2d[:, 7]
flytime[:, 2] = featime2d[:, 12]
flytime[:, 3] = featime2d[:, 15]
flytime[:, 4] = featime2d[:, 16]
flytime[:, 5] = featime2d[:, 18]

for i in range (flytime.shape[1]):
    for j in range(flytime.shape[0]):
        if flytime[j,i] >= 10.0:
            flytime[j,i] = np.mean(flytime[:,i])
        if flytime[j,i] <= 0.0:
            flytime[j, i] = 0.01


with open(rootfrom + "\\" + "weather5.pkl",'rb') as f:
    wea = pickle.load(f)
with open(rootfrom + "\\" + "spot.pkl",'rb') as f:
    spot = pickle.load(f)
minrow = min(len(flytime),len(wea),len(spot))
m = minrow
n = 1 + len(spot[0])+ flytime.shape[1]
input = np.zeros((m, n))
for i in range(m):
    for j in range(0,1):
        input[i, j] = wea[i]
    for j in range(1,1+len(spot[0])):
        input[i, j] = spot[i][j-1]
    for j in range(1+len(spot[0]),n):
        input[i, j] = flytime[i][j - 1 - len(spot[0])]

input2 = input

scaler1= MinMaxScaler(feature_range=(0, 1))
flytime = scaler1.fit_transform(flytime).squeeze()

scaler2= MinMaxScaler(feature_range=(0, 1))
input2 = scaler2.fit_transform(input2).squeeze()
conlution = np.zeros((5, 3))


for tsp in range(2,7):
    time_step = tsp
    batch_size = int(24*(60/sp))
    input_size = n
    hidden_size = 200
    layer_num = 2
    output_size = ((1,1))
    train_end = int((31+28+31)*24*(60/sp)) - time_step
    test_end = int(m) - time_step
    train_step = 3000
    test_step = (test_end - train_end) // batch_size


    def generate_data1(seq,seq2):
        x = []
        y = []
        for i in range(len(seq) - time_step):
            x.append([seq[i:i + time_step]])
            y.append([seq2[i + time_step]])
        xs = np.array(x, dtype = np.float32).squeeze()
        ys = np.array(y, dtype = np.float32).squeeze()
        return xs,ys


    def GRU_model(x,y,is_training):
        cell = tf.nn.rnn_cell.GRUCell(hidden_size)

        outputs, _ = tf.nn.dynamic_rnn(cell, x, dtype=tf.float32)
        output = outputs[:, -1, :]

        nonpredictions = tf.contrib.layers.fully_connected(output, output_size[0] * output_size[1], activation_fn=None)
        predictions = tf.nn.leaky_relu(nonpredictions, alpha=0.2, name=None)
        if not is_training:
            return predictions, None, None


        mse = [[] for i in range(output_size[0] * output_size[1])]
        for i in range(0, y.shape[1]):
            a = tf.reduce_mean(tf.square(y[:, i] - predictions[:, i]))
            mse[i].append(a)
        loss = tf.reduce_mean(mse)
        global_step = tf.Variable(0)
        LR = tf.train.exponential_decay(0.1, global_step, int(m / batch_size), 0.96, staircase=True)
        train_op = tf.contrib.layers.optimize_loss(loss, tf.train.get_global_step(), optimizer=opt, learning_rate=LR)
        return predictions, loss, train_op



    def train(sess,x,y,):
        ds = tf.data.Dataset.from_tensor_slices((x,y))
        ds = ds.repeat().shuffle(1000).batch(batch_size)
        x,y = ds.make_one_shot_iterator().get_next()

        with tf.variable_scope("grumodel",reuse = tf.AUTO_REUSE):
            _,train_op,lossor = GRU_model(x,y,True)

        sess.run(tf.global_variables_initializer())
        for i in range(train_step):
            _, loss = sess.run([train_op,lossor])
            if i % 100 ==0:
                print('lost: '+ str(i), loss)
    def test(sess,x,y):
        ds = tf.data.Dataset.from_tensor_slices((x, y))
        ds = ds.batch(batch_size)
        x, y = ds.make_one_shot_iterator().get_next()
        with tf.variable_scope("grumodel", reuse=tf.AUTO_REUSE):#为了使得测试的时候lstm调用的参数值是训练好的，必须设置reuse=True
            prediction, _, _ = GRU_model(x, [], False)
        for i in range(test_step):
            pred, ys = sess.run([prediction,y])#不能写y = 因为y和prediction被run以后就是ndarray,y这个变量已经存在了，类型是tensor，没法赋值，要能涵盖赋值，必须是同类型
            mse = np.zeros((ys.shape[0],ys.shape[1]))
            mae = np.zeros((ys.shape[0],ys.shape[1]))
            mape = np.zeros((ys.shape[0], ys.shape[1]))
            for k in range(0, ys.shape[1]):
                for l in range(0, ys.shape[0]):
                    a = np.square(ys[l, k] - pred[l, k])
                    c = np.abs(ys[l, k] - pred[l, k])
                    b = np.abs(ys[l, k] - pred[l, k]) / (ys[l, k])
                    mse[l,k] = a
                    mae[l,k] = c
                    mape[l,k] = b
            MSE = np.mean(mse,axis = 1)
            MAE = np.mean(mae,axis = 1)
            MAPE = np.mean(mape,axis = 1)
            ERROR = [MSE,MAE,MAPE]
            print("teststep:" ,i,ERROR)
            for e in range(0,3):
                errorcsv[i*batch_size:(i+1)*batch_size,e + 3*p] = ERROR[e]

    if __name__ == '__main__':
        sess = tf.Session()
        init = tf.global_variables_initializer()
        sess.run(init)
        x, y7 = generate_data1(input2, flytime)
        train_x = x[0:train_end]
        train_y7 = y7[0:train_end]
        train_y = np.zeros((train_y7.shape[0],1),dtype = np.float32)
        test_x = x[train_end:test_end]
        test_y7 = y7[train_end:test_end]
        test_y = np.zeros((test_y7.shape[0],1),dtype = np.float32)

        optimizer = ["SGD", "Adagrad", "Momentum"]
        errorcsv = np.zeros((test_step * batch_size, len(optimizer) * 3))
        for n in range(0,5):
            for i in range(train_y7.shape[0]):
                train_y[i,0] = train_y7[i,n]
            for i in range(test_y7.shape[0]):
                test_y[i,0] = test_y7[i, n]
            for p in range(0,len(optimizer)):
                opt = optimizer[p]
                b = os.path.exists(root + "\\" + sub + "\\" + opt)
                if b:
                    print("path exist")
                else:
                    os.makedirs(root + "\\" + sub + "\\" + opt)
                train(sess,train_x,train_y)
                test(sess,test_x,test_y)
            sixto24error = np.zeros((36*test_step, 9))
            for i in range(0,test_step):
                sixto24error[int(i*(24-6)*60/sp):int((i+1)*(24-6)*60/sp)] = errorcsv[int(24*60/sp*i+6*60/sp):int(24*60 / sp * (i+1))]
            for k in range(0,3):
                a = np.mean(sixto24error[:,k])
                b = np.mean(sixto24error[:,k+3])
                c = np.mean(sixto24error[:, k+6])
                s = (a + b + c)/3
                conlution[n, k] = s

            dt2 = pd.DataFrame(sixto24error, columns=[optimizer[0], '', '', optimizer[1], '', '', optimizer[2], '', ''])
            b = os.path.exists(root + "\\" + sub + "\\"+"error")
            if b:
                print("path exist")
            else:
                os.makedirs(root + "\\" + sub + "\\"+"error")
            dt2.to_csv(root + "\\" + sub + "\\"+"error" + "\\" + str(n) + "error" + str(time_step)+".csv")

        Writer = pd.ExcelWriter(track, engine = 'openpyxl')
        book = load_workbook(Writer.path)
        Writer.book = book
        dt1 = pd.DataFrame(conlution)
        dt1.to_excel(Writer, sheet_name = str(time_step))
        Writer.save()
        Writer.close()
