from sklearn.datasets import load_boston
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import SVR
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import numpy as np
import pickle
import pandas as pd
import  xlsxwriter
from openpyxl import load_workbook

# 1 准备数据
# 读取波士顿地区房价信息
# with open(r"D:\轨迹预测\ydian.pkl", 'rb') as f:
#      or_data = pickle.load(f)
# or_data = np.delete(or_data, 0, 0)
# data = or_data[0:1000, :]

MAEscalar = np.zeros((3, 10))
for t in range(0, 10):
    time_step = t*5 + 5

    root = r"D:\轨迹预测\prediction\SVR\nonscalar"
    with open(root + '\\' + 'test_x' + str(time_step) + ".pkl", 'rb') as f:
        test_x = pickle.load(f)
    with open(root + '\\' + 'test_y' + str(time_step) + ".pkl", 'rb') as f:
        test_y = pickle.load(f)

    pred = np.zeros((500, 3))
    for k in range(0, 3):
        x = test_x[1000:3000, :, k]
        # x = data[]
        y = test_y[1000:3000, k]

        x_train = x[500:len(x), :]
        y_train = y[500:len(y)]

        x_test = x[0:500, :]
        y_test = y[0:500]


        # 3 训练数据和测试数据进行标准化处理
        ss_x = StandardScaler()
        # ss_x = MinMaxScaler()
        x_train = ss_x.fit_transform(x_train)
        x_test = ss_x.transform(x_test)

        ss_y = StandardScaler()
        # ss_y = MinMaxScaler()
        y_train = ss_y.fit_transform(y_train.reshape(-1, 1))
        y_test = ss_y.transform(y_test.reshape(-1, 1))

        # 4.1 支持向量机模型进行学习和预测
        # 线性核函数配置支持向量机
        linear_svr = SVR(kernel="linear")
        # 训练
        linear_svr.fit(x_train, y_train)
        # 预测 保存预测结果
        linear_svr_y_predict = linear_svr.predict(x_test)


        # 多项式核函数配置支持向量机
        poly_svr = SVR(kernel="poly")
        # 训练
        poly_svr.fit(x_train, y_train)
        # 预测 保存预测结果
        poly_svr_y_predict = poly_svr.predict(x_test)



        # 高斯核函数
        rbf_svr = SVR(kernel="rbf")
        # 训练
        rbf_svr.fit(x_train, y_train)
        # 预测 保存预测结果
        rbf_svr_y_predict = rbf_svr.predict(x_test)


        #选择高斯
        pred[:, k] = ss_y.inverse_transform(linear_svr_y_predict[:])



        # 5 模型评估
        # 线性核函数 模型评估
        print("线性核函数支持向量机的默认评估值为：", linear_svr.score(x_test, y_test))
        print("线性核函数支持向量机的R_squared值为：", r2_score(y_test, linear_svr_y_predict))
        print("线性核函数支持向量机的均方误差为:", mean_squared_error(ss_y.inverse_transform(y_test),
                                                      ss_y.inverse_transform(linear_svr_y_predict)))
        print("线性核函数支持向量机的平均绝对误差为:", mean_absolute_error(ss_y.inverse_transform(y_test),
                                                         ss_y.inverse_transform(linear_svr_y_predict)))
        # 对多项式核函数模型评估
        print("对多项式核函数的默认评估值为：", poly_svr.score(x_test, y_test))
        print("对多项式核函数的R_squared值为：", r2_score(y_test, poly_svr_y_predict))
        print("对多项式核函数的均方误差为:", mean_squared_error(ss_y.inverse_transform(y_test),
                                                   ss_y.inverse_transform(poly_svr_y_predict)))
        print("对多项式核函数的平均绝对误差为:", mean_absolute_error(ss_y.inverse_transform(y_test),
                                                      ss_y.inverse_transform(poly_svr_y_predict)))

        # 高斯核函数评估
        print("对高斯核函数的默认评估值为：", rbf_svr.score(x_test, y_test))
        print("对高斯核函数的R_squared值为：", r2_score(y_test, rbf_svr_y_predict))
        print("对高斯核函数的均方误差为:", mean_squared_error(ss_y.inverse_transform(y_test),
                                                   ss_y.inverse_transform(rbf_svr_y_predict)))
        print("对高斯核函数的平均绝对误差为:", mean_absolute_error(ss_y.inverse_transform(y_test),
                                                      ss_y.inverse_transform(rbf_svr_y_predict)))

        MAEscalar[k, t] = mean_absolute_error(ss_y.inverse_transform(y_test),
                                                   ss_y.inverse_transform(rbf_svr_y_predict))


dt = pd.DataFrame(pred)
dt.to_csv(r"D:\轨迹预测\prediction\SVR\pred_y.csv")

dt = pd.DataFrame(MAEscalar)
dt.to_csv(r"D:\轨迹预测\prediction\SVR\MAE5_50.csv")

# track = r'D:\轨迹预测\conlution.xlsx'
# xls = xlsxwriter.Workbook(track)
# sheet = xls.add_worksheet('sheet1')
# xls.close()
# Writer = pd.ExcelWriter(track, engine='openpyxl')
# book = load_workbook(Writer.path)
# Writer.book = book
# dt1 = pd.DataFrame(MAEscalar)
# dt1.to_excel(Writer, sheet_name='SVR')
# Writer.save()